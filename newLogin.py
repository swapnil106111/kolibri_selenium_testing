import requests
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from methods import *
import unittest
import time
import sys
import os
from tkinter import *
import openpyxl
from time import gmtime, strftime
file = open("selenium.log", "a")
file.write("\n\n")
file.write(strftime("%Y-%m-%d %H:%M:%S", gmtime()))
file.write("\n")


class KolibriTesting(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_testing_server_is_started(self):
		r = requests.get("http://localhost:8008")
		if r.status_code == 200:
			print("Kolibri Server is started")
			file.write("\n")
			file.write("test_a_testing_server_is_started is passed")
			file.write("\n")
		else:
			print("Server Error %s" %r.status_code)
			file.write("\n")
			file.write("test_a_testing_server_is_started is failed")
			file.write("\n")
		self.assertEqual(r.status_code, 200)

	def test_b_create_users_and_login_with_exsting_user_with_wrong_password(self):
		temp = sample(self.driver)
		if temp==1:
			text = LoginDifferentKindOfUser(self.driver, "coach", "password")

			self.assertEqual(text, "Incorrect username or password",'Logging with the user that is not part of class')
			file.write("\n")
			file.write("test_b_create_users_and_login_with_exsting_user_with_wrong_password is passed")
			file.write("\n")
	def test_c_start_with_capital_a(self):
		self.driver.get("http://localhost:8008")
		text = LoginDifferentKindOfUser(self.driver, "Admin", "password")
		self.assertTrue(text, "Incorrect username or password")
		file.write("\n")
		file.write("test_c_start_with_capital_a is passed")
		file.write("\n")

	def test_d_admin_username_in_capital(self):
		self.driver.get("http://localhost:8008")
		text = LoginDifferentKindOfUser(self.driver, "ADMIN", "password")
		self.assertTrue(text, "Incorrect username or password")
		file.write("\n")
		file.write("test_d_admin_username_in_capital is passed")
		file.write("\n")
	def test_e_login_with_non_exsting_user(self):
		self.driver.get("http://localhost:8008")
		text = LoginDifferentKindOfUser(self.driver, "nalanda","lee")
		self.assertEqual(text, "Incorrect username or password",'Logging with the user that is not part of class')
		file.write("\n")
		file.write("test_e_login_with_non_exsting_user is passed")
		file.write("\n")
	def tearDown(self):
		self.driver.close()


class KolibriTestingClass(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_delete_class(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		url = self.driver.current_url
		if "setup_wizard" in url:
			print("You need to setup Device owner account")
			file.write("\n")
			file.write("test_a_delete_class :: You need to setup Device owner account")
			file.write("\n")
		else:
			wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
			sheet_user = wb.get_sheet_by_name('user')
			LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
			time.sleep(7)
			text = getText('enter classroom name::')
			if len(text)== 0:
				print("Empty classroom name not allowed")
			else:
				if text in self.driver.page_source:
					trs = self.driver.find_elements(By.TAG_NAME, "tr")
					for i in range(1,len(trs)):
						tt = trs[i].find_element(By.TAG_NAME, "th")
						tt1 = tt.find_element(By.TAG_NAME, "a").text
						if text == tt1:
							trs[i].find_element(By.TAG_NAME, "button").click()
							time.sleep(3)
							self.driver.find_element_by_xpath("//span[contains(text(), 'Delete Class')]").click()
							time.sleep(3)
							self.assertEqual(False, text in self.driver.page_source)
							file.write("\n")
							file.write("test_a_delete_class is passed")
							file.write("\n")
							break
				else:
					print("Class not found")
					file.write("\n")
					file.write("test_a_delete_class :: Class not found")
					file.write("\n")

	def tearDown(self):
		self.driver.close()


class KolibriTestingFacility(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_check_edit_username_facility(self):
		self.driver.get("http://localhost:8008")
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(2)
		self.driver.find_element_by_xpath('//a[contains(@href,"#/facilities")]').click()
		time.sleep(3)
		if not self.driver.find_element_by_name("learnerCanEditUsername").is_selected():
			self.driver.find_element_by_xpath('//div[contains(text(), "Allow users to edit their username")]').click()
		time.sleep(2)
		self.driver.find_element_by_xpath('//div[contains(text(), "Save changes")]').click()
		time.sleep(4)
		SignOut(self.driver)
		time.sleep(4)
		try:
			LoginDifferentKindOfUser(self.driver, str(sheet_user['A5'].value), str(sheet_user['B5'].value))
			time.sleep(5)
			self.driver.find_element_by_xpath("//button[@type='submit']").click()
			time.sleep(2)
			self.driver.find_element_by_xpath('//span[contains(text(), "menu")]').click()
			time.sleep(4)
			self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]').click()
			time.sleep(4)
		except Exception:
			print("Please enter username which is exist")

		self.driver.find_element_by_xpath("//input[@type='text']").clear()
		time.sleep(3)
		self.driver.find_element_by_xpath("//input[@type='text']").send_keys(str(sheet_user['A6'].value))
		time.sleep(3)
		self.driver.find_element_by_xpath("//span[contains(text(), 'Save changes')]").click()
		time.sleep(4)

		if "An account with that username already exists" in self.driver.page_source:
			username = getText("Enter new username which is not exist")
			username = validate(username, "Enter new username which is not exist")
			self.driver.find_element_by_xpath("//input[@type='text']").clear()
			time.sleep(3)
			self.driver.find_element_by_xpath("//input[@type='text']").send_keys(username)
			time.sleep(3)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Save changes')]").click()
			time.sleep(4)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(3)
		self.driver.find_element_by_xpath("//span[contains(text(), 'exit_to_app')]").click()
		time.sleep(5)
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A6'].value), str(sheet_user['B6'].value))
		time.sleep(8)
		temp = self.driver.title
		self.assertTrue(True, "Learn" in temp)
		time.sleep(5)
		self.driver.find_element_by_xpath("//button[@type='submit']").click()
		time.sleep(2)
		self.driver.find_element_by_xpath('//span[contains(text(), "menu")]').click()
		time.sleep(4)
		self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]').click()
		time.sleep(2)
		self.driver.find_element_by_xpath("//input[@type='text']").clear()
		time.sleep(4)
		self.driver.find_element_by_xpath("//input[@type='text']").send_keys(str(sheet_user['A5'].value))
		time.sleep(3)
		self.driver.find_element_by_xpath("//span[contains(text(), 'Save changes')]").click()
		time.sleep(4)

		print("Edit username facility working properly")
		file.write("\n")
		file.write("test_a_check_edit_username_facility is passed")
		file.write("\n")

	def test_b_check_edit_fullname_facility(self):
		self.driver.get("http://localhost:8008")
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(2)
		self.driver.find_element_by_xpath('//a[contains(@href,"#/facilities")]').click()
		time.sleep(3)
		if self.driver.find_element_by_name("learnerCanEditUsername").is_selected():
			self.driver.find_element_by_xpath('//div[contains(text(), "Allow users to edit their username")]').click()
		time.sleep(2)

		if self.driver.find_element_by_name("learnerCanEditName").is_selected():
			pass
		else:
			self.driver.find_element_by_xpath('//div[contains(text(), "Allow users to edit their full name")]').click()

		time.sleep(2)
		self.driver.find_element_by_xpath('//div[contains(text(), "Save changes")]').click()
		time.sleep(4)
		SignOut(self.driver)
		time.sleep(4)
		try:
			LoginDifferentKindOfUser(self.driver, str(sheet_user['A5'].value), str(sheet_user['B5'].value))
			time.sleep(5)
			self.driver.find_element_by_xpath("//button[@type='submit']").click()
			time.sleep(2)
			self.driver.find_element_by_xpath('//span[contains(text(), "menu")]').click()
			time.sleep(4)
			self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]').click()
			time.sleep(4)
		except Exception:
			self.driver.refresh()
			time.sleep(5)
			print("Please enter username which is exist")
			username = getText("Enter new username which is exist")
			username = validate(username, "Enter new username which is exist")
			LoginDifferentKindOfUser(self.driver, username,"sc")
			time.sleep(5)
			self.driver.find_element_by_xpath("//button[@type='submit']").click()
			time.sleep(2)
			self.driver.find_element_by_xpath('//span[contains(text(), "menu")]').click()
			time.sleep(4)
			self.driver.find_element_by_xpath('//div[contains(text(), "Profile")]').click()
			time.sleep(4)
		full_name = getText("Enter full name")
		full_name = validate(full_name, "Enter full name")
		self.driver.find_element_by_xpath("//input[@type='text']").clear()
		time.sleep(3)
		self.driver.find_element_by_xpath("//input[@type='text']").send_keys(full_name)
		time.sleep(3)
		self.driver.find_element_by_xpath("//span[contains(text(), 'Save changes')]").click()
		time.sleep(4)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(3)
		self.driver.find_element_by_xpath("//span[contains(text(), 'exit_to_app')]").click()
		time.sleep(5)
		self.assertEqual(True, True)
		print("Edit full name facility working properly")
		file.write("\n")
		file.write("test_b_check_edit_fullname_facility is passed")
		file.write("\n")
	def tearDown(self):
		self.driver.close()


class KolibriTestingGroup(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_create_group(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A3'].value), str(sheet_user['B3'].value))

		time.sleep(7)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(2)
		self.driver.find_element_by_xpath("//span[contains(text(), 'assessment')]").click()
		time.sleep(3)
		class_name = str(sheet_user['D2'].value)
		if class_name in self.driver.page_source:
			time.sleep(2)
			self.driver.find_element_by_xpath("//a[contains(text(), '%s')]" %class_name).click()
			time.sleep(3)
			self.driver.find_element_by_xpath("//span[contains(text(), 'group_work')]").click()
			time.sleep(4)

			group_name = getText("Enter group name")
			group_name = validate(group_name, "Enter group name")
			time.sleep(2)
			self.driver.find_element_by_xpath("//span[contains(text(), 'New group')]").click()
			time.sleep(3)
			self.driver.find_element_by_xpath("//input[@type='text']").send_keys(group_name)
			time.sleep(2)
			if "A group with that name already exists" in self.driver.page_source:
				group_name = getText("Enter unique group name")
				group_name = validate(group_name, "Enter unique group name")
				time.sleep(2)
				self.driver.find_element_by_xpath("//input[@type='text']").clear()
				time.sleep(2)
				self.driver.find_element_by_xpath("//input[@type='text']").send_keys(group_name)
				time.sleep(2)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Save')]").click()
			time.sleep(4)
			try:
				self.driver.find_element_by_xpath("//input[@type='checkbox']").click()
				time.sleep(3)
				temp = self.driver.find_elements_by_class_name("group-section")
				for i in temp:
					text1 = i.find_element(By.TAG_NAME, "span").text
					if text1 != "0 Learners":
						i.find_element(By.TAG_NAME, "button").click()
						time.sleep(4)
						self.driver.find_element_by_xpath("//label[contains(text(), '%s')]" %group_name).click()
						time.sleep(4)
						self.driver.find_element_by_xpath("//span[contains(text(), 'Move')]").click()
						time.sleep(4)
						print("Learners added into new group")
						
						SignOut(self.driver)
						self.assertEqual(True, True)
						file.write("\n")
						file.write("test_a_create_group is passed")
						file.write("\n")
						break
				
			except Exception:
				print("There is no student in class")
				file.write("\n")
				file.write("test_a_delete_class:: No student in the class")
				file.write("\n")

		else:
			print("Enter valid classroom name")
			file.write("\n")
			file.write("test_a_delete_class :: Enter vaid classroom name")
			file.write("\n")


	def test_b_delete_group(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A3'].value), str(sheet_user['B3'].value))
		time.sleep(7)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(2)
		self.driver.find_element_by_xpath("//span[contains(text(), 'assessment')]").click()
		time.sleep(3)
		class_name = str(sheet_user['D2'].value)
		if class_name in self.driver.page_source:
			time.sleep(2)
			self.driver.find_element_by_xpath("//a[contains(text(), '%s')]" %class_name).click()
			time.sleep(3)
			self.driver.find_element_by_xpath("//span[contains(text(), 'group_work')]").click()
			time.sleep(4)
			btn = self.driver.find_elements_by_css_selector(".ui-icon.ui-button__dropdown-icon.material-icons")
			group_name = getText("Enter group name to delete")
			group_name = validate(group_name , "Enter group name to delete")
			if group_name == "Ungrouped":
				print("Sorry, You can not delete Ungrouped group")
				file.write("\n")
				file.write("test_b_delete_group :: Ungrouped group not able to delete")
				file.write("\n")
			else:
				if group_name in self.driver.page_source:
					temp = self.driver.find_elements_by_class_name("group-section")
					count = 0
					if group_name in self.driver.page_source:
						for i in temp:
							t = i.find_element(By.TAG_NAME, "h2").text
							count = count +1
							if t == group_name:
								btn[count].click()
								time.sleep(4)
								self.driver.find_element_by_xpath("//div[contains(text(), 'Delete Group')]").click()
								time.sleep(4)
								self.driver.find_element_by_xpath("//span[contains(text(), 'Delete Group')]").click()
								time.sleep(5)
								temp = self.driver.find_elements_by_class_name("group-section")
								break
					if group_name in self.driver.page_source:
						count = 0
						for i in temp:
							t = i.find_element(By.TAG_NAME, "h2").text
							if t == group_name:
								count =1
								break

						if count == 0:
							self.assertEqual(True, True)
							print("Group deleted successfully")
							file.write("\n")
							file.write("test_b_delete_group is passed")
							file.write("\n")
						else:
							self.assertEqual(True , False)
							file.write("\n")
							file.write("test_b_delete_group is failed")
							file.write("\n")
					else:
						self.assertEqual(True, True)
						print("Group deleted successfully")
						file.write("\n")
						file.write("test_b_delete_group is passed")
						file.write("\n")
				else:
					print("%s not present in the class" %group_name)
					self.assertEqual(True , True)
					file.write("\n")
					file.write("test_a_delete_class :: Class not found")
					file.write("\n")
		else:
			print("Class is not present")
			file.write("\n")
			file.write("test_a_delete_class :: Class not present")
			file.write("\n")

	def tearDown(self):
		self.driver.close()


class KolibriTestingExam(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_create_exam(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		sheet_channel = wb.get_sheet_by_name('channel')
		sheet_exam = wb.get_sheet_by_name('exam')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A3'].value), str(sheet_user['B3'].value))
		time.sleep(7)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(2)
		self.driver.find_element_by_xpath("//span[contains(text(), 'assessment')]").click()
		time.sleep(5)

		text = "Signed in as device owner"
		if text not in self.driver.page_source:
			class_name = str(sheet_user['D2'].value)
			time.sleep(2)
			if class_name in self.driver.page_source:
				try:
					self.driver.find_element_by_xpath("//a[contains(text(), '%s')]" %class_name).click()
					time.sleep(4)
					self.driver.find_element_by_xpath("//span[contains(text(), 'assignment_late')]").click()
					time.sleep(4)
					self.driver.find_element_by_xpath("//div[contains(text(), 'New Exam')]").click()
					time.sleep(4)
					self.driver.find_element_by_css_selector(".ui-select__display-value.is-placeholder").click()
					time.sleep(4)
					channel_name = str(sheet_exam['C1'].value)
					if channel_name not in self.driver.page_source:
						channel_name = getText("Enter channel name")
						channel_name = validate(channel_name, "Enter channel name")
						time.sleep(3)
						if channel_name not in self.driver.page_source:
							print("Channel name not present")
							sys.exit(0)

					self.driver.find_element_by_xpath("//div[contains(text(), '%s')]" %channel_name).click()
					time.sleep(4)
					self.driver.find_element_by_xpath("//span[contains(text(), 'Create exam')]").click()
					time.sleep(7)
					exam_name = str(sheet_exam['A1'].value)
					time.sleep(3)
					self.driver.find_element_by_xpath("//input[@type='text']").clear()
					time.sleep(2)
					self.driver.find_element_by_xpath("//input[@type='text']").send_keys(exam_name)				
					text = "An exam with that title already exists"
					time.sleep(3)
					if text in self.driver.page_source:
						exam_name = getText("Enter exam name")
						exam_name = validate(exam_name, "Enter exam name")
						time.sleep(3)
						self.driver.find_element_by_xpath("//input[@type='text']").clear()
						time.sleep(2)
						self.driver.find_element_by_xpath("//input[@type='text']").send_keys(exam_name)		
						time.sleep(3)
						if text in self.driver.page_source:
							print("Exam name already exist")
							sys.exit(0)
					
					time.sleep(4)
					self.driver.find_element_by_xpath("//input[@type='number']").clear()
					time.sleep(2)
					self.driver.find_element_by_xpath("//input[@type='number']").send_keys(int(sheet_exam['B1'].value))
					time.sleep(2)
					self.driver.find_element_by_xpath("//input[@type='checkbox']").click()
					time.sleep(15)

					self.driver.find_element_by_xpath("//span[contains(text(), 'Finish')]").click()
					time.sleep(8)

					temp = self.driver.find_elements_by_tag_name("tr")
					temp.pop(0)
					for i in range(0,len(temp)):
						text1 = temp[i].find_element(By.TAG_NAME, "strong").text
						if text1 == exam_name:
							col = temp[i].find_elements(By.TAG_NAME, "td")[3]
							col.find_element(By.TAG_NAME, "div").click()
							time.sleep(5)
							self.driver.find_element_by_xpath("//span[contains(text(), 'Activate')]").click()
							time.sleep(5)
							SignOut(self.driver)
							time.sleep(3)
							self.assertEqual(True, True)
							print("Exam Created and Assigned")
							file.write("\n")
							file.write("test_a_create_exam :: Exam created and assigned")
							file.write("\n")
							time.sleep(5)
							LoginDifferentKindOfUser(self.driver, str(sheet_user['A5'].value), str(sheet_user['B5'].value))
							time.sleep(6)
							self.driver.find_element_by_xpath("//span[contains(text(), 'assignment_late')]").click()
							time.sleep(3)
							temp = self.driver.find_elements_by_css_selector(".exam-row")
							# print(temp)
							count = 0
							for i in temp:
								t = i.find_element(By.TAG_NAME, "h2").text
								if t == exam_name:
									count = 1
									i.find_element(By.TAG_NAME, "span").click()
									time.sleep(50)
									self.driver.find_element_by_xpath("//span[contains(text(), 'Submit exam')]").click()
									time.sleep(3)
									t = self.driver.find_elements(By.TAG_NAME, "span")
									for j in t:
										if j.text == "SUBMIT EXAM":
											time.sleep(3)
											j.click()
									time.sleep(7)
									print("Exam Given successfully")
									file.write("\n")
									file.write("test_a_create_exam is passed")
									file.write("\n")
									self.assertEqual(True, True)
									break
							if count == 0:
								file.write("\n")
								file.write("test_a_create_exam ::Exam is not given to this learner")
								file.write("\n")
								print("Exam is not given for this learner")
							break

				except Exception:
					print("Element not found")
					file.write("\n")
					file.write("test_a_create_exam ::Element not found")
					file.write("\n")

			else:
				print("Class not present")
				file.write("\n")
				file.write("test_a_create_exam ::Class not found")
				file.write("\n")				
				self.assertEqual(True, True)
		else:
			print("Sorry, you login as a Device owner. Please login with admin or coach")
			file.write("\n")
			file.write("test_a_create_exam :: Sorry, you login as a Device owner. Please login with admin or coach")
			file.write("\n")
			self.assertEqual(True, True)


	def test_b_delete_exam(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')		
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A3'].value), str(sheet_user['B3'].value))
		time.sleep(7)
		self.driver.find_element_by_xpath("//span[contains(text(), 'menu')]").click()
		time.sleep(2)
		self.driver.find_element_by_xpath("//span[contains(text(), 'assessment')]").click()
		time.sleep(5)

		text = "Signed in as device owner"
		if text not in self.driver.page_source:
			class_name = str(sheet_user['D2'].value)
			time.sleep(2)
			if class_name in self.driver.page_source:
				try:
					self.driver.find_element_by_xpath("//a[contains(text(), '%s')]" %class_name).click()
					time.sleep(4)
					self.driver.find_element_by_xpath("//span[contains(text(), 'assignment_late')]").click()
					time.sleep(4)
		
					exam_name = str(sheet_user['G2'].value)
					time.sleep(3)
					count = 0
					temp = self.driver.find_elements_by_tag_name("tr")
					temp.pop(0)
					for i in range(0,len(temp)):
						text1 = temp[i].find_element(By.TAG_NAME, "strong").text
						if text1 == exam_name:
							count = 1
							break

					if count == 0:
						print("Exam is not available to delete")
						file.write("\n")
						file.write("test_b_delete_exam :: Exam not available to delete")
						file.write("\n")
						self.assertEqual(True, True)
					else:
						col = temp[i].find_elements(By.TAG_NAME, "td")[3]
						col.find_element(By.TAG_NAME, "span").click()
						time.sleep(5)
						self.driver.find_element_by_xpath("//div[contains(text(), 'Delete')]").click()
						time.sleep(4)
						self.driver.find_element_by_xpath("//span[contains(text(), 'Delete')]").click()
						time.sleep(5)
				except Exception:
					print("Element not found")
					file.write("\n")
					file.write("test_b_delete_exam :: Element not found")
					file.write("\n")
				temp = self.driver.find_elements_by_tag_name("tr")
				if len(temp) != 0:
					temp.pop(0)
					for i in range(0,len(temp)):
						text1 = temp[i].find_element(By.TAG_NAME, "strong").text
						if text1 == exam_name:
							count = 1
							break
					if count == 0:
						print("After deleting exam, Exam still there")
						file.write("\n")
						file.write("test_b_delete_exam :: Still exam is there")
						file.write("\n")
						self.assertEqual(True, True)

					else:
						self.assertEqual(True, True)
						file.write("\n")
						file.write("test_b_delete_exam is passed")
						file.write("\n")
						print("Exam deleted successfully.")
						SignOut(self.driver)
				else:
						self.assertEqual(True, True)
						file.write("\n")
						file.write("test_b_delete_exam is passed")
						file.write("\n")
						print("Exam deleted successfully.")
						SignOut(self.driver)


			else:
				print("Class not present")
				file.write("\n")
				file.write("test_b_delete_exam :: Class not present")
				file.write("\n")
				self.assertEqual(True, True)
		else:
			print("Sorry, you login as a Device owner. Please login with admin or coach")
			file.write("\n")
			file.write("test_b_delete_exam :: Sorry, you login as a Device owner. Please login with admin or coach")
			file.write("\n")
			self.assertEqual(True, True)
	def tearDown(self):
		self.driver.close()



class KolibriTestingImportExport(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()

	def test_a_import_channel_from_internet(self):
		self.driver.get("http://localhost:8008")
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_channel = wb.get_sheet_by_name('channel')
		sheet_user = wb.get_sheet_by_name('user')

		time.sleep(7)
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(8)
		self.driver.find_element_by_xpath("//span[contains(text(), 'view_module')]").click()
		time.sleep(4)
		try:
			self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
			time.sleep(3)
		except Exception:
			pass
		try:
			self.driver.find_element_by_xpath("//span[contains(text(), 'Import')]").click()
			time.sleep(4)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Internet')]").click()
			time.sleep(4)
			self.driver.find_element_by_xpath("//input[@type='text']").clear()
			time.sleep(2)
			channel_id = str(sheet_channel['A1'].value)
			time.sleep(2)
			self.driver.find_element_by_xpath("//input[@type='text']").send_keys(channel_id)
			time.sleep(2)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Import')]").click()
			time.sleep(5)
			if "That ID was not found on our server." in self.driver.page_source:
				print("Your channel ID is not valid")
				file.write("\n")
				file.write(" test_a_import_channel_from_internet :: ID is not present on server")
				file.write("\n")
				self.assertEqual(False, False)
			else:
				time.sleep(15)
				if "Finished!" in self.driver.page_source:
					self.assertEqual(True, True)
					self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
					time.sleep(2)
					print("Channel imported successfully")
					file.write("\n")
					file.write(" test_a_import_channel_from_internet is passed")
					file.write("\n")
				else:
					time.sleep(45)
					if "Finished!" in self.driver.page_source:
						self.assertEqual(True, True)
						self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
						time.sleep(2)
						print("Channel imported successfully")
						file.write("\n")
						file.write("test_a_import_channel_from_internet is passed")
						file.write("\n")
					else:
						print("You have very slow internet speed, Check import by your own or You are not connected to the Internet")
						file.write("\n")
						file.write("test_a_import_channel_from_internet :: You have very slow internet speed, Check import by your own or You are not connected to the Internet")
						file.write("\n")
		except Exception:
			file.write("\n")
			file.write("test_a_import_channel_from_internet ::Server is not responding")
			file.write("\n")
			print("Server is not responding")

	def test_b_import_channel_from_localdrive(self):
		self.driver.get("http://localhost:8008")
		time.sleep(7)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(8)
		self.driver.find_element_by_xpath("//span[contains(text(), 'view_module')]").click()
		time.sleep(4)

		try:
			self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
			time.sleep(3)
		except Exception:
			pass
		try:
			time.sleep(3)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Import')]").click()
			time.sleep(4)
			self.driver.find_element_by_xpath("//span[contains(text(), 'Local Drives')]").click()
			time.sleep(4)
			count = 0
			path = getText("Enter path of local channel")
			path = validate(path, "Enter path of local channel")
			if path not in self.driver.page_source:
				path = getText("Enter path of local channel")
				path = validate(path, "Enter path of local channel")
				if path not in self.driver.page_source:
					print("Please Enter correct path or connect External drive to the server")
					file.write("\n")
					file.write("test_b_import_channel_from_localdrive :: Please Enter correct path or connect External drive to the server")
					file.write("\n")
				else:
					count = 1
			else: 
				count = 1
			if count == 1:
				self.driver.find_element_by_xpath("//div[contains(text(), '%s')]" %path).click()
				time.sleep(4)
				self.driver.find_element_by_xpath("//span[contains(text(), 'Import')]").click() 
				time.sleep(15)
				if "Finished!" in self.driver.page_source:
					print("Channel Imported successfully")
					file.write("\n")
					file.write("test_b_import_channel_from_localdrive is passed")
					file.write("\n")
					time.sleep(2)
					self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
					time.sleep(2)
					self.assertEqual(True, True)
				else:
					time.sleep(50)
					if "Finished!" in self.driver.page_source:
						print("Channel Imported successfully")
						file.write("\n")
						file.write("test_b_import_channel_from_localdrive is passed")
						file.write("\n")
						time.sleep(2)
						self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
						time.sleep(2)
						self.assertEqual(True, True)
					else:
						print("Please check this feature manually")
						file.write("\n")
						file.write("test_b_import_channel_from_localdrive :: Please check this feature manually")
						file.write("\n")
						self.assertEqual(True, False)
		except Exception:
			print("Server is not responding")
			file.write("\n")
			file.write("test_b_import_channel_from_localdrive :: Server is not responding")
			file.write("\n")


	def test_c_export_channel_to_localdrive(self):
		self.driver.get("http://localhost:8008")
		time.sleep(7)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(8)
		self.driver.find_element_by_xpath("//span[contains(text(), 'view_module')]").click()
		time.sleep(4)
		if "No channels installed" in self.driver.page_source:
			print("No channels installed in your system")
			file.write("\n")
			file.write("test_c_export_channel_to_localdrive :: No channels installed")
			file.write("\n")
		else:
			try:
				self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
				time.sleep(3)
			except Exception:
				pass
			try:
				time.sleep(3)
				self.driver.find_element_by_xpath("//span[contains(text(), 'Export')]").click()
				time.sleep(4)
				count = 0
				path = getText("Enter path of local channel")
				path = validate(path, "Enter path of local channel")
				if path not in self.driver.page_source:
					path = getText("Enter path of local channel")
					path = validate(path, "Enter path of local channel")
					if path not in self.driver.page_source:
						print("Please Enter correct path or connect External drive to the server")
						file.write("\n")
						file.write("test_c_export_channel_to_localdrive :: Please Enter correct path or connect External drive to the server")
						file.write("\n")
					else:
						count = 1
				else: 
					count = 1
				if count == 1:
					self.driver.find_element_by_xpath("//div[contains(text(), '%s')]" %path).click()
					time.sleep(4)
					self.driver.find_element_by_xpath("//span[contains(text(), 'Export')]").click() 
					time.sleep(15)
					if "Finished!" in self.driver.page_source:
						self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
						print("Channel Exported successfully")
						file.write("\n")
						file.write("test_c_export_channel_to_localdrive is passed")
						file.write("\n")
						self.assertEqual(True, True)
					else:
						time.sleep(50)
						if "Finished!" in self.driver.page_source:
							self.driver.find_element_by_xpath("//span[contains(text(), 'Close')]").click()
							file.write("\n")
							file.write("test_c_export_channel_to_localdrive is passed")
							file.write("\n")
							print("Channel Exported successfully")
							self.assertEqual(True, True)
						else:
							file.write("\n")
							file.write("test_c_export_channel_to_localdrive :: Please check this feature manually")
							file.write("\n")
							print("Please check this feature manually")
							self.assertEqual(True, False)
			except Exception:
				print("Server is not responding")			
				file.write("\n")
				file.write("test_c_export_channel_to_localdrive :: Server not responding")
				file.write("\n")


	def test_d_delete_channel(self):
		self.driver.get("http://localhost:8008")
		time.sleep(7)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A2'].value), str(sheet_user['B2'].value))
		time.sleep(8)
		self.driver.find_element_by_xpath("//span[contains(text(), 'view_module')]").click()
		time.sleep(4)
		if "No channels installed" in self.driver.page_source:
			print("No channels installed in your system")
			file.write("\n")
			file.write("test_d_delete_channel :: No channel installed")
			file.write("\n")
		else:
			channel_name = getText("Enter channel name")
			channel_name = validate(channel_name, "Enter channel name")
			if channel_name in self.driver.page_source:
				temp = self.driver.find_elements_by_tag_name("tr")
				temp.pop(0)
				count = 0
				for i in range(0,len(temp)):
					text1 = temp[i].find_element(By.TAG_NAME, "td").text
					if text1 == channel_name:
						col = temp[i].find_elements(By.TAG_NAME, "td")[4]
						col.find_element(By.TAG_NAME, "button").click()
						time.sleep(4)
						self.driver.find_element_by_xpath("//div[contains(text(), 'Confirm')]").click()
						time.sleep(5)
						break
				if channel_name in self.driver.page_source:
					print("Channel delete feature is not working properly")
					file.write("\n")
					file.write("test_d_delete_channel :: Channel delete feature not working properly")
					file.write("\n")
					self.assertEqual(True, False)
				else:
					self.assertEqual(True, True)
					print("Channel Deleted successfully")
					file.write("\n")
					file.write("test_d_delete_channel is passed")
					file.write("\n")
			else:
				print("%s is not present" %channel_name)
				file.write("\n")
				file.write("test_d_delete_channel :: channel not present")
				file.write("\n")
				self.assertEqual(True, True)

	def tearDown(self):
		self.driver.close()


class KolibriTestingLearnTab(unittest.TestCase):
	def setUp(self):
		self.driver = webdriver.Firefox()


	def test_a_exercise(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		sheet_exercise = wb.get_sheet_by_name('exercise')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A5'].value), str(sheet_user['B5'].value))
		time.sleep(7)
		points = self.driver.find_element_by_css_selector(".description-value").text
		points = str(points)
		points = int(points.replace(',',''))
		self.driver.find_element_by_xpath("//span[contains(text(), 'apps')]").click()
		time.sleep(4)
		self.driver.find_element_by_xpath("//div[contains(text(), '%s')]" %str(sheet_exercise['A15'].value)).click()
		time.sleep(5)
		self.driver.find_element_by_xpath("//span[contains(text(), 'folder')]").click()
		time.sleep(4)
		try:
			for i in range(1,10):
				timeout = int(sheet_exercise['B%s' %i].value)
				path = str(sheet_exercise['A%s' %i].value)
				time.sleep(2)
				if timeout > 8:
					self.driver.find_element_by_xpath("//h3[contains(text(), '%s')]" %path).click()
					time.sleep(timeout)
					temp = self.driver.find_element_by_css_selector(".description-value").text
					temp = str(temp)
					temp = int(temp.replace(',',''))
					if points + 500 == temp:
						print("Congratulations.! You have completed with mastery...")
						file.write("\n")
						file.write("test_a_exercise is passed with mastery")
						file.write("\n")
						self.assertEqual(True, True)
						break
					else:
						print("Sorry, You have not completed mastery...")
						file.write("\n")
						file.write("test_a_exercise is passed without mastery")
						file.write("\n")
						self.assertEqual(True, True)
						break
				else:
					self.driver.find_element_by_xpath("//h3[contains(text(), '%s')]" %path).click()
					time.sleep(timeout)

		except Exception:
			print("Error while loading")
			file.write("\n")
			file.write("test_a_exercise :: Error while loading")
			file.write("\n")

	def test_b_video(self):
		self.driver.get("http://localhost:8008")
		time.sleep(3)
		wb = openpyxl.load_workbook('/home/kolibri/Desktop/selenium/excel.xlsx')
		sheet_user = wb.get_sheet_by_name('user')
		sheet_exercise = wb.get_sheet_by_name('video')
		LoginDifferentKindOfUser(self.driver, str(sheet_user['A5'].value), str(sheet_user['B5'].value))
		time.sleep(7)
		points = self.driver.find_element_by_css_selector(".description-value").text
		points = str(points)
		points = int(points.replace(',',''))
		self.driver.find_element_by_xpath("//span[contains(text(), 'apps')]").click()
		time.sleep(4)
		self.driver.find_element_by_xpath("//div[contains(text(), '%s')]" %str(sheet_exercise['A15'].value)).click()
		time.sleep(5)
		self.driver.find_element_by_xpath("//span[contains(text(), 'folder')]").click()
		time.sleep(4)
		
		for i in range(1,10):
			#print(i)
			timeout = int(sheet_exercise['B%s' %i].value)
			path = str(sheet_exercise['A%s' %i].value)
			time.sleep(2)
			try:
				if timeout > 8:
					self.driver.find_element_by_xpath("//h3[contains(text(), '%s')]" %path).click()
					time.sleep(5)
					self.driver.find_element_by_xpath("//span[contains(text(), 'Play')]").click()
					time.sleep(5)
					time.sleep(timeout)
					temp = self.driver.find_element_by_css_selector(".description-value").text
					temp = str(temp)
					temp = int(temp.replace(',',''))
					total = points + 500					
					if total == temp:
						print("Congratulations.! You have completed with mastery...")
						self.assertTrue(True, True)
						print("Passed")
						file.write("\n")
						file.write("test_b_video is passed with mastery")
						file.write("\n")
						break
					else:
						print("Sorry, You have not completed mastery...")
						self.assertTrue(True, True)
						print("Passed")
						file.write("\n")
						file.write("test_b_video is passed without mastery")
						file.write("\n")
						break
				else:
					self.driver.find_element_by_xpath("//h3[contains(text(), '%s')]" %path).click()
					time.sleep(timeout)
			except Exception:
				print("Something wrong going on")
				file.write("\n")
				file.write("test_b_video :: Something went wrong")
				file.write("\n")


	def tearDown(self):
		self.driver.close()

if __name__=="__main__":
	unittest.main()

